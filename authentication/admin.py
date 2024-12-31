from django.contrib import admin
from . import models
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils import timezone

@admin.register(models.User)
class UserAdmin(admin.ModelAdmin):
    list_display = ('uniqueIdentifier', 'mobile' , 'referal','create_at')
    search_fields = ('uniqueIdentifier', 'mobile' , 'referal')
    list_filter = ('uniqueIdentifier', 'mobile' , 'referal')
    list_per_page = 100
    ordering = ['uniqueIdentifier']
    fieldsets = (
        ('اطلاعات اصلی', {
            'fields': ('uniqueIdentifier', 'mobile')
        }),
        ('اطلاعات معرف', {
            'fields': ('referal',)
        }),
        ('اطلاعات نوع', {
            'fields': ('type',)
        }),
        ('اطلاعات وضعیت', {
            'fields': ('status',)
        }),
        ('اطلاعات نماینده', {
            'fields': ('agent',)
        }),
        ('اطلاعات درخواست ها', {
            'fields': ('attempts',)
        }),
        ('اطلاعات قفل', {
            'fields': ('lock_until',)
        }),

    )
    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات کاربران"

        headers = [
            'شناسه یکتا',
            'موبایل',
            'ایمیل',
            'نوع',
            'وضعیت',
            'معرف',
            'نماینده'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                obj.uniqueIdentifier,
                obj.mobile,
                obj.email,
                obj.type,
                obj.status,
                obj.referal,
                obj.agent
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=users_{timezone.now().date()}.xlsx'

        wb.save(response)
        return response

    export_as_excel.short_description = "خروجی اکسل از موارد انتخاب شده"

@admin.register(models.Otp)
class OtpAdmin(admin.ModelAdmin):
    list_display = ('mobile', 'code', 'expire')
    search_fields = ('mobile', 'code', 'expire')
    list_filter = ('mobile', 'code', 'expire')
    list_per_page = 25
    ordering = ['mobile']
    fieldsets = (
        ('اطلاعات اصلی', {
            'fields': ('mobile', 'code', 'expire')
        }),
    )
    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات کد ها"

        headers = [
            'موبایل',
            'کد',
            'تاریخ انقضا'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                obj.mobile,
                obj.code,
                obj.expire.replace(tzinfo=None) if obj.expire else None
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=otp_{timezone.now().date()}.xlsx'

        wb.save(response)
        return response

    export_as_excel.short_description = "خروجی اکسل از موارد انتخاب شده"

@admin.register(models.Admin)
class AdminAdmin(admin.ModelAdmin):
    list_display = ('uniqueIdentifier', 'mobile', 'email' , 'lastName' , 'firstName')
    search_fields = ('uniqueIdentifier', 'mobile', 'email' , 'lastName' , 'firstName')
    list_filter = ('uniqueIdentifier', 'mobile', 'email' , 'lastName' , 'firstName')
    list_per_page = 25
    ordering = ['uniqueIdentifier']
    actions = ['export_as_excel']
    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات ادمین ها"

        headers = [
            'شناسه یکتا',
            'موبایل',
            'ایمیل',
            'نام',
            'نام خانوادگی',
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                obj.uniqueIdentifier,
                obj.mobile,
                obj.email,
                obj.firstName,
                obj.lastName,
                obj.attempts,
                obj.lock_until.replace(tzinfo=None) if obj.lock_until else None
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=private_persons_{timezone.now().date()}.xlsx'

        wb.save(response)
        return response

    export_as_excel.short_description = "خروجی اکسل از موارد انتخاب شده"
    fieldsets = (
        ('اطلاعات اصلی', {
            'fields': ('uniqueIdentifier', 'mobile', 'email' , 'lastName' , 'firstName')
        }),
        ('اطلاعات درخواست ها', {
            'fields': ('attempts',)
        }),
        ('اطلاعات قفل', {
            'fields': ('lock_until',)
        })

    )


@admin.register(models.privatePerson)
class privatePersonAdmin(admin.ModelAdmin):
    list_display = ('firstName', 'lastName', 'fatherName', 'birthDate', 'gender', 'shNumber')
    list_filter = ('gender',)
    search_fields = ('firstName', 'lastName', 'shNumber', 'serial')
    readonly_fields = ('user',)
    list_per_page = 25
    actions = ['export_as_excel']
    ordering = ['lastName', 'firstName']

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات اشخاص"

        headers = [
            'نام',
            'نام خانوادگی',
            'نام پدر',
            'تاریخ تولد',
            'جنسیت',
            'محل تولد',
            'محل صدور',
            'سری شناسنامه',
            'سریال',
            'شماره شناسنامه'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                obj.firstName,
                obj.lastName,
                obj.fatherName,
                obj.birthDate,
                obj.gender,
                obj.placeOfBirth,
                obj.placeOfIssue,
                obj.seriSh,
                obj.serial,
                obj.shNumber
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=private_persons_{timezone.now().date()}.xlsx'

        wb.save(response)
        return response

    export_as_excel.short_description = "خروجی اکسل از موارد انتخاب شده"

    fieldsets = (
        ('اطلاعات شخصی', {
            'fields': (
                'user',
                ('firstName', 'lastName'),
                'fatherName',
                'gender',
                'birthDate',
            )
        }),
        ('اطلاعات شناسنامه', {
            'fields': (
                'placeOfBirth',
                'placeOfIssue',
                ('seriSh', 'seriShChar'),
                'serial',
                'shNumber',
            )
        }),
        ('امضا', {
            'fields': ('signatureFile',)
        })
    )

@admin.register(models.tradingCodes)
class tradingCodesAdmin(admin.ModelAdmin):
    list_display = ('get_user_identifier', 'code')
    search_fields = ('user__uniqueIdentifier', 'code')
    list_filter = ('code', 'type')
    list_per_page = 25
    ordering = ['user']
    
    def get_user_identifier(self, obj):
        return obj.user.uniqueIdentifier if obj.user else None
    get_user_identifier.short_description = 'شناسه کاربر'
    get_user_identifier.admin_order_field = 'user__uniqueIdentifier'

    fieldsets = (
        ('اطلاعات اصلی', {
            'fields': ('user', 'code')
        }),
        ('اطلاعات جزئی', {
            'fields': (
                'firstPart',
                'secondPart',
                'thirdPart',
                'type'
            )
        }),
    )
    actions = ['export_as_excel']
    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات کد ها"

        headers = [
            'نام کاربری',
            'کد',
            'قسمت اول',
            'قسمت دوم',
            'قسمت سوم',
            'نوع'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                str(obj.user.uniqueIdentifier),
                obj.code,
                obj.firstPart,
                obj.secondPart,
                obj.thirdPart,
                obj.type
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=trading_codes_{timezone.now().date()}.xlsx'

        wb.save(response)
        return response

    export_as_excel.short_description = "خروجی اکسل از موارد انتخاب شده"

@admin.register(models.jobInfo)
class jobInfoAdmin(admin.ModelAdmin):
    list_display = ('get_user_identifier', 'companyName', 'job', 'position')
    search_fields = ('user__uniqueIdentifier', 'companyName', 'job', 'position')
    list_filter = ('companyName', 'job', 'position')
    list_per_page = 25
    ordering = ['user']
    
    def get_user_identifier(self, obj):
        return obj.user.uniqueIdentifier if obj.user else None
    get_user_identifier.short_description = 'شناسه کاربر'
    get_user_identifier.admin_order_field = 'user__uniqueIdentifier'

    fieldsets = (
        ('اطلاعات اصلی', {
            'fields': ('user', 'companyName', 'job', 'position')
        }),
        ('اطلاعات جزئی', {
            'fields': (
                'companyAddress', 
                'companyCityPrefix',
                'companyEmail', 
                'companyFax',
                'companyFaxPrefix',
                'companyPhone',
                'companyPostalCode',
                'companyWebSite',
                'employmentDate',
                'jobDescription'
            )
        }),
    )

    actions = ['export_as_excel']
    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات کارها"

        headers = [
            'نام کاربری',
            'نام شرکت',
            'شغل',
            'موقعیت',
            'آدرس',
            'پیش شهر',
            'ایمیل',
            'فکس',
            'پیش شهر فکس',
            'تاریخ استخدام',
            'شغل',
            'توضیحات'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                str(obj.user.uniqueIdentifier),
                obj.companyName,
                obj.job,
                obj.position,
                obj.companyAddress,
                obj.companyCityPrefix,
                obj.companyEmail,
                obj.companyFax,
                obj.companyFaxPrefix,
                obj.employmentDate,
                obj.jobDescription,
                obj.job,
                obj.position    
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=job_info_{timezone.now().date()}.xlsx'

        wb.save(response)
        return response

    export_as_excel.short_description = "خروجی اکسل از موارد انتخاب شده"

@admin.register(models.financialInfo)
class financialInfoAdmin(admin.ModelAdmin):
    list_display = ('get_user_identifier', 'companyPurpose', 'financialBrokers', 'referenceRateCompany', 'sExchangeTransaction', 'tradingKnowledgeLevel', 'transactionLevel')
    search_fields = ('user__uniqueIdentifier', 'companyPurpose', 'financialBrokers', 'referenceRateCompany', 'sExchangeTransaction', 'tradingKnowledgeLevel', 'transactionLevel')
    list_filter = ('companyPurpose', 'financialBrokers', 'referenceRateCompany', 'sExchangeTransaction', 'tradingKnowledgeLevel', 'transactionLevel')
    list_per_page = 25
    ordering = ['user']
    
    def get_user_identifier(self, obj):
        return obj.user.uniqueIdentifier if obj.user else None
    get_user_identifier.short_description = 'شناسه کاربر'
    get_user_identifier.admin_order_field = 'user__uniqueIdentifier'

    fieldsets = (
        ('اطلاعات اصلی', {
            'fields': ('user', 'companyPurpose', 'financialBrokers', 'referenceRateCompany', 'sExchangeTransaction', 'tradingKnowledgeLevel', 'transactionLevel')
        }),
        ('اطلاعات جزئی', {
            'fields': (
                'assetsValue',
                'cExchangeTransaction',
                'inComingAverage',
                'outExchangeTransaction',
                'rate',
                'rateDate'
            )
        }),
    )

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات مالی"

        headers = [
            'نام کاربری',
            'نام شرکت',
            'شغل',
            'موقعیت',
            'آدرس',
            'پیش شهر',
            'ایمیل',
            'فکس',
            'پیش شهر فکس',
            'تاریخ استخدام',
            'شغل',
            'توضیحات'
        ]

        ws.append(headers)

        for obj in queryset:
            row = [
                str(obj.user.uniqueIdentifier),
                obj.companyPurpose,
                obj.financialBrokers,
                obj.referenceRateCompany,
                obj.sExchangeTransaction,
                obj.tradingKnowledgeLevel,
                obj.transactionLevel
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=financial_info_{timezone.now().date()}.xlsx'

        wb.save(response)
        return response

    export_as_excel.short_description = "خروجی اکسل از موارد انتخاب شده"

@admin.register(models.addresses)
class addressesAdmin(admin.ModelAdmin):
    list_display = ('get_user_identifier', 'city', 'country', 'postalCode', 'province', 'remnantAddress', 'section', 'tel', 'website')
    search_fields = ('user__uniqueIdentifier', 'city', 'country', 'postalCode', 'province', 'remnantAddress', 'section', 'tel', 'website')
    list_filter = ('user__uniqueIdentifier', 'city', 'country', 'postalCode', 'province', 'remnantAddress', 'section', 'tel', 'website')
    list_per_page = 25
    ordering = ['user']
    actions = ['export_as_excel']
    def get_user_identifier(self, obj):
        return obj.user.uniqueIdentifier
    get_user_identifier.short_description = 'شناسه کاربر'
    get_user_identifier.admin_order_field = 'user__uniqueIdentifier'

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات سدرس‌ها"

        # تعریف ستون‌ها با عرض مناسب
        columns = [
            ('نام کاربری', 20),
            ('شهر', 15),
            ('کشور', 15),
            ('کد پستی', 12),
            ('استان', 15),
            ('آدرس', 40),
            ('بخش', 15),
            ('تلفن', 15),
            ('وب‌سایت', 30)
        ]

        # اضافه کردن هدرها و تنظیم عرض ستون‌ها
        for col, (header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            ws.column_dimensions[cell.column_letter].width = width

        # اضافه کردن داده‌ها
        for row, obj in enumerate(queryset, 2):
            data = [
                str(obj.user.uniqueIdentifier),
                obj.city or '',
                obj.country or '',
                obj.postalCode or '',
                obj.province or '',
                obj.remnantAddress or '',
                obj.section or '',
                obj.tel or '',
                obj.website or ''
            ]
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=addresses_{timezone.now().date()}.xlsx'
        wb.save(response)
        return response

    export_as_excel.short_description = "دریافت خروجی اکسل"

    fieldsets = (
        ('اطلاعات کاربر', {
            'fields': ('user',)
        }),
        ('اطلاعات آدرس', {
            'fields': (
                ('city', 'province', 'country'),
                ('postalCode', 'section'),
                'remnantAddress',
            )
        }),
        ('اطلاعات تماس', {
            'fields': (
                ('tel', 'website'),
            )
        })
    )

@admin.register(models.accounts)
class accountsAdmin(admin.ModelAdmin):
    list_display = ('get_user_identifier', 'accountNumber', 'bank', 'sheba', 'branchName', 'type', 'isDefault')
    search_fields = ('user__uniqueIdentifier', 'accountNumber', 'bank', 'sheba', 'branchName')
    list_filter = ('bank', 'type', 'isDefault', 'branchCity')
    list_per_page = 25
    ordering = ['user', 'bank']
    
    def get_user_identifier(self, obj):
        return obj.user.uniqueIdentifier
    get_user_identifier.short_description = 'شناسه کاربر'
    get_user_identifier.admin_order_field = 'user__uniqueIdentifier'

    fieldsets = (
        ('اطلاعات کاربر', {
            'fields': ('user',)
        }),
        ('اطلاعات حساب', {
            'fields': (
                ('accountNumber', 'sheba'),
                ('bank', 'type'),
                'isDefault',
            )
        }),
        ('اطلاعات شعبه', {
            'fields': (
                ('branchName', 'branchCode'),
                'branchCity',
            )
        }),
        ('سایر اطلاعات', {
            'fields': ('modifiedDate',),
            'classes': ('collapse',)
        })
    )

    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات حساب‌های بانکی"

        # تعریف ستون‌ها با عرض مناسب
        columns = [
            ('شناسه کاربر', 20),
            ('شماره حساب', 20),
            ('شماره شبا', 30),
            ('نام بانک', 15),
            ('نوع حساب', 15),
            ('پیش‌فرض', 10),
            ('نام شعبه', 20),
            ('کد شعبه', 10),
            ('شهر شعبه', 15),
            ('تاریخ آخرین تغییر', 20)
        ]

        # اضافه کردن هدرها و تنظیم عرض ستون‌ها
        for col, (header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            ws.column_dimensions[cell.column_letter].width = width

        # اضافه کردن داده‌ها
        for row, obj in enumerate(queryset, 2):
            data = [
                str(obj.user.uniqueIdentifier),
                obj.accountNumber,
                obj.sheba,
                obj.bank,
                obj.type or '',
                obj.isDefault or '',
                obj.branchName or '',
                obj.branchCode or '',
                obj.branchCity or '',
                obj.modifiedDate or ''
            ]
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=bank_accounts_{timezone.now().date()}.xlsx'
        wb.save(response)
        return response

    export_as_excel.short_description = "دریافت خروجی اکسل"

@admin.register(models.LegalPerson)
class LegalPersonAdmin(admin.ModelAdmin):
    list_display = ('companyName', 'economicCode', 'registerNumber', 'get_user_identifier')
    search_fields = ('companyName', 'economicCode', 'registerNumber', 'user__uniqueIdentifier')
    list_filter = ('citizenshipCountry', 'legalPersonTypeCategory')
    list_per_page = 25
    ordering = ['companyName']

    def get_user_identifier(self, obj):
        return obj.user.uniqueIdentifier
    get_user_identifier.short_description = 'شناسه کاربر'
    get_user_identifier.admin_order_field = 'user__uniqueIdentifier'

    fieldsets = (
        ('اطلاعات کاربر', {
            'fields': ('user',)
        }),
        ('اطلاعات شرکت', {
            'fields': (
                'companyName',
                ('economicCode', 'registerNumber'),
                ('registerPlace', 'registerDate'),
                'citizenshipCountry',
            )
        }),
        ('نوع شخص حقوقی', {
            'fields': (
                'legalPersonTypeCategory',
                'legalPersonTypeSubCategory',
            )
        }),
        ('اطلاعات مدارک', {
            'fields': (
                'evidenceReleaseCompany',
                ('evidenceReleaseDate', 'evidenceExpirationDate'),
            )
        }),
    )

    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات اشخاص حقوقی"

        headers = [
            'شناسه کاربر',
            'نام شرکت',
            'کد اقتصادی',
            'شماره ثبت',
            'محل ثبت',
            'تاریخ ثبت',
            'کشور تابعیت',
            'نوع شخص حقوقی',
            'زیرمجموعه نوع شخص حقوقی',
            'شرکت صادرکننده مدرک',
            'تاریخ صدور مدرک',
            'تاریخ انقضای مدرک'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                str(obj.user.uniqueIdentifier),
                obj.companyName or '',
                obj.economicCode or '',
                obj.registerNumber or '',
                obj.registerPlace or '',
                obj.registerDate or '',
                obj.citizenshipCountry or '',
                obj.legalPersonTypeCategory or '',
                obj.legalPersonTypeSubCategory or '',
                obj.evidenceReleaseCompany or '',
                obj.evidenceReleaseDate or '',
                obj.evidenceExpirationDate or ''
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=legal_persons_{timezone.now().date()}.xlsx'
        wb.save(response)
        return response

    export_as_excel.short_description = "دریافت خروجی اکسل"

@admin.register(models.legalPersonStakeholders)
class LegalPersonStakeholdersAdmin(admin.ModelAdmin):
    list_display = ('firstName', 'lastName')
    search_fields = ('firstName', 'lastName')
    list_filter = ('firstName', 'lastName')  # حذف legalPerson
    list_per_page = 25
    ordering = ['firstName', 'lastName']


    fieldsets = (
        ('اطلاعات شرکت', {
            'fields': ('legalPerson',)
        }),
        ('اطلاعات فردی', {
            'fields': (
                ('firstName', 'lastName'),
            )
        }),
        ('اطلاعات تکمیلی', {
            'fields': (
                'startDate',
                'endDate',
                'signatureFile'
            )
        }),
    )

    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات ذینفعان"

        headers = [
            'نام شرکت',
            'نام',
            'نام خانوادگی', 
            'تاریخ شروع',
            'تاریخ پایان'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                obj.legalPerson.companyName,
                obj.firstName,
                obj.lastName,
                obj.startDate,
                obj.endDate
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=stakeholders_{timezone.now().date()}.xlsx'

        wb.save(response)
        return response

    export_as_excel.short_description = "دریافت خروجی اکسل"

@admin.register(models.legalPersonShareholders) 
class legalPersonShareholdersAdmin(admin.ModelAdmin):
    list_display = ('firstName', 'lastName', 'uniqueIdentifier', 'positionType', 'percentageVotingRight')
    search_fields = ('firstName', 'lastName', 'uniqueIdentifier', 'positionType')
    list_filter = ('positionType',)
    list_per_page = 25
    ordering = ['lastName', 'firstName']
   
    fieldsets = (
        ('اطلاعات کاربر', {
            'fields': ('user',)
        }),
        ('اطلاعات شخصی', {
            'fields': (
                ('firstName', 'lastName'),
                'uniqueIdentifier',
                'positionType',
                'percentageVotingRight',
            )
        }),
        ('اطلاعات تماس', {
            'fields': (
                'postalCode',
                'address',
            )
        }),
    )

    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات سهامداران"

        headers = [
            'شناسه کاربر',
            'نام',
            'نام خانوادگی',
            'شناسه یکتا',
            'نوع موقعیت',
            'درصد حق رای',
            'کد پستی',
            'آدرس'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                str(obj.user.uniqueIdentifier),
                obj.firstName or '',
                obj.lastName or '',
                obj.uniqueIdentifier or '',
                obj.positionType or '',
                obj.percentageVotingRight or '',
                obj.postalCode or '',
                obj.address or ''
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=shareholders_{timezone.now().date()}.xlsx'
        wb.save(response)
        return response

    export_as_excel.short_description = "دریافت خروجی اکسل"

admin.site.register(models.BlacklistedToken)
admin.site.register(models.Reagent)

